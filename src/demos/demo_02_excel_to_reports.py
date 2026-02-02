"""
Demo 02: Excel Data to Template-Based Reports

This demonstration shows how to:
1. Load data from a CSV file
2. Calculate simple metrics (Month-over-Month, Year-over-Year growth)
3. Read an Excel template
4. Populate the template with calculated data
5. Save the populated report

This is an instructor-led demo. Students will implement this workflow
themselves in Exercise 2.

Run from repository root:
    python src/demos/demo_02_excel_to_reports.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os


def main():
    """Run the Excel-to-Reports demonstration"""

    print("="*60)
    print("DEMO: Excel Data to Template-Based Reports")
    print("="*60)
    print()

    # STEP 1: Load data from CSV
    print("STEP 1: Loading data from CSV")
    print("-" * 40)

    csv_path = 'data/sample_monthly_data.csv'

    try:
        df = pd.read_csv(csv_path)
        print(f"✓ Loaded {len(df)} indicators from {csv_path}")
        print(f"  Columns: {', '.join(df.columns.tolist())}")
        print(f"  Indicators: {', '.join(df['indicator_code'].tolist()[:3])}...")
        print()
    except FileNotFoundError:
        print(f"✗ Error: {csv_path} not found")
        print("  Make sure you're running from the repository root!")
        return

    # STEP 2: Calculate metrics
    print("STEP 2: Calculating metrics")
    print("-" * 40)

    # Add calculated columns for growth rates
    df['mom_change_pct'] = ((df['value'] - df['previous_month_value']) / df['previous_month_value'] * 100).round(2)
    df['yoy_change_pct'] = ((df['value'] - df['previous_year_value']) / df['previous_year_value'] * 100).round(2)

    print("✓ Calculated Month-over-Month (MoM) growth rates")
    print("✓ Calculated Year-over-Year (YoY) growth rates")
    print()

    # Show sample calculations
    print("Sample calculations:")
    sample = df.iloc[0]
    print(f"  {sample['indicator_name']}:")
    print(f"    Current: {sample['value']:,.0f} {sample['unit']}")
    print(f"    Previous Month: {sample['previous_month_value']:,.0f} {sample['unit']}")
    print(f"    MoM Change: {sample['mom_change_pct']:.2f}%")
    print(f"    YoY Change: {sample['yoy_change_pct']:.2f}%")
    print()

    # STEP 3: Load Excel template
    print("STEP 3: Loading Excel template")
    print("-" * 40)

    template_path = 'templates/monthly_report_template.xlsx'

    try:
        workbook = openpyxl.load_workbook(template_path)
        print(f"✓ Loaded template from {template_path}")
        print(f"  Sheets: {', '.join(workbook.sheetnames)}")
        print()
    except FileNotFoundError:
        print(f"✗ Error: {template_path} not found")
        return

    # STEP 4: Populate the template
    print("STEP 4: Populating template with data")
    print("-" * 40)

    # Get the Summary sheet
    sheet = workbook['Summary']

    # Update report metadata
    reporting_period = df['period'].iloc[0]
    sheet['B2'] = f"Monthly Economic Indicators Report - {reporting_period}"
    sheet['B3'] = reporting_period

    print(f"✓ Set report title and period: {reporting_period}")

    # Populate indicator data (starting at row 5)
    start_row = 5

    for idx, row in df.iterrows():
        current_row = start_row + idx

        # Column B: Indicator name
        sheet[f'B{current_row}'] = row['indicator_name']

        # Column C: Current value
        sheet[f'C{current_row}'] = row['value']

        # Column D: Previous month value
        sheet[f'D{current_row}'] = row['previous_month_value']

        # Column E: Month-over-month change percentage
        sheet[f'E{current_row}'] = row['mom_change_pct'] / 100  # Convert to decimal for Excel percentage formatting

    print(f"✓ Populated {len(df)} indicators into the template")
    print()

    # STEP 5: Generate insights (simple rule-based approach)
    print("STEP 5: Generating insights")
    print("-" * 40)

    insights = []

    # Find top performers (highest YoY growth)
    top_growth = df.nlargest(3, 'yoy_change_pct')
    for idx, row in top_growth.iterrows():
        insights.append(
            f"{row['indicator_name']} grew {row['yoy_change_pct']:.1f}% year-over-year"
        )

    # Find declining indicators
    declining = df[df['yoy_change_pct'] < 0]
    for idx, row in declining.iterrows():
        insights.append(
            f"{row['indicator_name']} decreased {abs(row['yoy_change_pct']):.1f}% year-over-year"
        )

    # Add insights to template (starting at row 15)
    insight_row = 15
    for i, insight in enumerate(insights[:5]):  # Limit to 5 insights
        sheet[f'B{insight_row + i}'] = f"• {insight}"

    print(f"✓ Generated {len(insights)} insights")
    for insight in insights[:3]:
        print(f"  • {insight}")
    if len(insights) > 3:
        print(f"  ... and {len(insights) - 3} more")
    print()

    # STEP 6: Save the populated report
    print("STEP 6: Saving populated report")
    print("-" * 40)

    # Create output directory if it doesn't exist
    os.makedirs('output', exist_ok=True)

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f'output/monthly_report_DEMO_{timestamp}.xlsx'

    workbook.save(output_path)
    print(f"✓ Report saved to: {output_path}")
    print()

    # Summary
    print("="*60)
    print("DEMO COMPLETE!")
    print("="*60)
    print()
    print("What you learned:")
    print("  1. Load data from CSV using pandas")
    print("  2. Calculate metrics (MoM, YoY growth rates)")
    print("  3. Load an Excel template with openpyxl")
    print("  4. Populate template cells with calculated data")
    print("  5. Generate rule-based insights")
    print("  6. Save the populated report")
    print()
    print("Next steps:")
    print(f"  • Open the generated report: {output_path}")
    print("  • Try Exercise 2: exercises/exercise_02_reports.py")
    print()


if __name__ == '__main__':
    main()
